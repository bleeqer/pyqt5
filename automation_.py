from PyQt5 import QtCore, QtGui, QtWidgets
from openpyxl import load_workbook
import datetime
from PyQt5 import Qt
import os


def importing_recrods(workbook):
    wb = load_workbook(workbook)
    sheetNames = wb.sheetnames[:]
    removeList = []
    records = []
    for id in range(len(sheetNames)):
	    if len(str(sheetNames[id])) > 2:
	        removeList.append(sheetNames[id]) 

    for sheet in removeList:
        sheetNames.remove(sheet)

    for sheetName in sheetNames:
        ws = wb[sheetName]


        for i in range(0, 6):
            val = ws.cell(36-i, 2).value
            if val == None:
                continue
            data = ["{}".format(sheetName), val[3:]]
            records.append(data)
            
    wb.close()
    return records



class Ui_MainWindow(object):
    def __init__(self):
        self.matched = {} 
        self.workbookList = ['files/기계식.xlsx', 'files/유압식.xlsx', 'files/기타설비.xlsx']
        self.columns = {'A': 1, 'B': 2, 'C': 3}
        self.machineName = ''
        self.reversedRecords = records[::-1]
        self.idx = 0
        self.year = 0
        self.month = 0
        self.strDate = ''        
        self.dateFormat = '%Y%m%d'
        self.editedRec = ''
        self.matching_sig = 0
        self.deletedLists = []
        self.deletedEdited = []
        self.machineNameList = []
        self.matchingText = ''
        self.listOrEdit = 3
        # self.deletedItem = QWidgets.QListWidgetItem()
        


    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1100, 850)
        MainWindow.setCursor(QtGui.QCursor(QtCore.Qt.ArrowCursor))
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        
        self.shortcut = QtWidgets.QShortcut(QtGui.QKeySequence("Ctrl+Z"), self.centralwidget)
        self.shortcut.activated.connect(self.rollback)
        
        self.gridLayoutWidget = QtWidgets.QWidget(self.centralwidget)
        self.gridLayoutWidget.setGeometry(QtCore.QRect(590, 90, 482, 471))
        self.gridLayoutWidget.setObjectName("gridLayoutWidget")
        self.gridLayout = QtWidgets.QGridLayout(self.gridLayoutWidget)
        self.gridLayout.setContentsMargins(0, 0, 0, 0)
        self.gridLayout.setObjectName("gridLayout")
        self.MP2002 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP2002.setAutoFillBackground(False)
        self.MP2002.setObjectName("MP2002")
        self.gridLayout.addWidget(self.MP2002, 0, 7, 1, 1)
        self.MP2002.clicked.connect(lambda: self.matching("MP2002"))
        self.MP1103 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP1103.setAutoFillBackground(False)
        self.MP1103.setObjectName("MP1103")
        self.gridLayout.addWidget(self.MP1103, 0, 4, 1, 1)
        self.MP1103.clicked.connect(lambda: self.matching("MP1103"))
        self.adding = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.adding.setAutoFillBackground(False)
        self.adding.setObjectName("Adding")
        self.gridLayout.addWidget(self.adding, 11, 7, 1, 1)
        self.adding.clicked.connect(self.adding_details)
        
        self.deleting = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.deleting.setAutoFillBackground(False)
        self.deleting.setObjectName("Deleting")
        self.gridLayout.addWidget(self.deleting, 12, 7, 1, 1)
        self.deleting.clicked.connect(self.deleting_details)
        self.MP1101 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP1101.setAutoFillBackground(False)
        self.MP1101.setObjectName("MP1101")
        self.gridLayout.addWidget(self.MP1101, 0, 2, 1, 1)
        self.MP1101.clicked.connect(lambda: self.matching("MP1101"))
        self.MP1102 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP1102.setAutoFillBackground(False)
        self.MP1102.setObjectName("MP1102")
        self.gridLayout.addWidget(self.MP1102, 0, 3, 1, 1)
        self.MP1102.clicked.connect(lambda: self.matching("MP1102"))
        self.MP2001 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP2001.setAutoFillBackground(False)
        self.MP2001.setObjectName("MP2001")
        self.gridLayout.addWidget(self.MP2001, 0, 6, 1, 1)
        self.MP2001.clicked.connect(lambda: self.matching("MP2001"))
        self.MP1104 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP1104.setAutoFillBackground(False)
        self.MP1104.setStyleSheet("")
        self.MP1104.setObjectName("MP1104")
        self.gridLayout.addWidget(self.MP1104, 0, 5, 1, 1)
        self.MP1104.clicked.connect(lambda: self.matching("MP1104"))
        self.MP4005 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP4005.setAutoFillBackground(False)
        self.MP4005.setObjectName("MP4005")
        self.gridLayout.addWidget(self.MP4005, 1, 7, 1, 1)
        self.MP4005.clicked.connect(lambda: self.matching("MP4005"))
        self.MP4004 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP4004.setAutoFillBackground(False)
        self.MP4004.setObjectName("MP4004")
        self.gridLayout.addWidget(self.MP4004, 1, 6, 1, 1)
        self.MP4004.clicked.connect(lambda: self.matching("MP4004"))
        self.MP4003 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP4003.setAutoFillBackground(False)
        self.MP4003.setObjectName("MP4003")
        self.gridLayout.addWidget(self.MP4003, 1, 5, 1, 1)
        self.MP4003.clicked.connect(lambda: self.matching("MP4003"))
        self.MP4002 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP4002.setAutoFillBackground(False)
        self.MP4002.setObjectName("MP4002")
        self.gridLayout.addWidget(self.MP4002, 1, 4, 1, 1)
        self.MP4002.clicked.connect(lambda: self.matching("MP4002"))
        self.MP4001 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP4001.setAutoFillBackground(False)
        self.MP4001.setObjectName("MP4001")
        self.gridLayout.addWidget(self.MP4001, 1, 3, 1, 1)
        self.MP4001.clicked.connect(lambda: self.matching("MP4001"))
        self.MP2501 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP2501.setAutoFillBackground(False)
        self.MP2501.setObjectName("MP2501")
        self.gridLayout.addWidget(self.MP2501, 1, 2, 1, 1)
        self.MP2501.clicked.connect(lambda: self.matching("MP2501"))
        self.MP6304 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP6304.setAutoFillBackground(False)
        self.MP6304.setObjectName("MP6304")
        self.gridLayout.addWidget(self.MP6304, 2, 7, 1, 1)
        self.MP6304.clicked.connect(lambda: self.matching("MP6304"))
        self.MP6303 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP6303.setAutoFillBackground(False)
        self.MP6303.setObjectName("MP6303")
        self.gridLayout.addWidget(self.MP6303, 2, 6, 1, 1)
        self.MP6303.clicked.connect(lambda: self.matching("MP6303"))
        self.MP6302 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP6302.setAutoFillBackground(False)
        self.MP6302.setObjectName("MP6302")
        self.gridLayout.addWidget(self.MP6302, 2, 5, 1, 1)
        self.MP6302.clicked.connect(lambda: self.matching("MP6302"))
        self.MP6301 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP6301.setAutoFillBackground(False)
        self.MP6301.setObjectName("MP6301")
        self.gridLayout.addWidget(self.MP6301, 2, 4, 1, 1)
        self.MP6301.clicked.connect(lambda: self.matching("MP6301"))
        self.MP600 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP600.setAutoFillBackground(False)
        self.MP600.setObjectName("MP600")
        self.gridLayout.addWidget(self.MP600, 2, 3, 1, 1)
        self.MP600.clicked.connect(lambda: self.matching("MP600"))
        self.MP4006 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP4006.setAutoFillBackground(False)
        self.MP4006.setObjectName("MP4006")
        self.gridLayout.addWidget(self.MP4006, 2, 2, 1, 1)
        self.MP4006.clicked.connect(lambda: self.matching("MP4006"))
        self.MP1000 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP1000.setAutoFillBackground(False)
        self.MP1000.setObjectName("MP1000")
        self.gridLayout.addWidget(self.MP1000, 4, 4, 1, 1)
        self.MP1000.clicked.connect(lambda: self.matching("MP1000"))
        self.MP650 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP650.setAutoFillBackground(False)
        self.MP650.setObjectName("MP650")
        self.gridLayout.addWidget(self.MP650, 4, 2, 1, 1)
        self.MP650.clicked.connect(lambda: self.matching("MP650"))
        self.MP800 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.MP800.setAutoFillBackground(False)
        self.MP800.setObjectName("MP800")
        self.gridLayout.addWidget(self.MP800, 4, 3, 1, 1)
        self.MP800.clicked.connect(lambda: self.matching("MP800"))
        self.HP201 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP201.setObjectName("HP201")
        self.gridLayout.addWidget(self.HP201, 5, 2, 1, 1)
        self.HP201.clicked.connect(lambda: self.matching("HP201"))
        self.HP202 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP202.setObjectName("HP202")
        self.gridLayout.addWidget(self.HP202, 5, 3, 1, 1)
        self.HP202.clicked.connect(lambda: self.matching("HP202"))
        self.HP203 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP203.setObjectName("HP203")
        self.gridLayout.addWidget(self.HP203, 5, 4, 1, 1)
        self.HP203.clicked.connect(lambda: self.matching("HP203"))
        self.HP204 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP204.setObjectName("HP204")
        self.gridLayout.addWidget(self.HP204, 5, 5, 1, 1)
        self.HP204.clicked.connect(lambda: self.matching("HP204"))
        self.HP205 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP205.setObjectName("HP205")
        self.gridLayout.addWidget(self.HP205, 5, 6, 1, 1)
        self.HP205.clicked.connect(lambda: self.matching("HP205"))
        self.HP206 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP206.setObjectName("HP206")
        self.gridLayout.addWidget(self.HP206, 5, 7, 1, 1)
        self.HP206.clicked.connect(lambda: self.matching("HP206"))
        self.HP207 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP207.setObjectName("HP207")
        self.gridLayout.addWidget(self.HP207, 9, 2, 1, 1)
        self.HP207.clicked.connect(lambda: self.matching("HP207"))
        self.HP208 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP208.setObjectName("HP208")
        self.gridLayout.addWidget(self.HP208, 9, 3, 1, 1)
        self.HP208.clicked.connect(lambda: self.matching("HP208"))
        self.SB1 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.SB1.setObjectName("sb1")
        self.gridLayout.addWidget(self.SB1, 10, 7, 1, 1)
        self.SB1.clicked.connect(lambda: self.matching("SB1"))
        self.SB3 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.SB3.setObjectName("sb3")
        self.gridLayout.addWidget(self.SB3, 11, 3, 1, 1)
        self.SB3.clicked.connect(lambda: self.matching("SB3"))
        self.HT1 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HT1.setObjectName("HT1")
        self.gridLayout.addWidget(self.HT1, 11, 4, 1, 1)
        self.HT1.clicked.connect(lambda: self.matching("HT1"))
        self.HP209 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HP209.setObjectName("HP209")
        self.gridLayout.addWidget(self.HP209, 9, 4, 1, 1)
        self.HP209.clicked.connect(lambda: self.matching("HP209"))
        self.CS1 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.CS1.setObjectName("CS1")
        self.gridLayout.addWidget(self.CS1, 10, 2, 1, 1)
        self.CS1.clicked.connect(lambda: self.matching("CS1"))
        self.CS2 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.CS2.setObjectName("CS2")
        self.gridLayout.addWidget(self.CS2, 10, 3, 1, 1)
        self.CS2.clicked.connect(lambda: self.matching("CS2"))
        self.CS3 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.CS3.setObjectName("cs3")
        self.gridLayout.addWidget(self.CS3, 10, 4, 1, 1)
        self.CS3.clicked.connect(lambda: self.matching("CS3"))
        
        self.CS4 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.CS4.setObjectName("cs4")
        self.gridLayout.addWidget(self.CS4, 10, 5, 1, 1)
        self.CS4.clicked.connect(lambda: self.matching("CS4"))
        
        self.CS5 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.CS5.setObjectName("cs5")
        self.gridLayout.addWidget(self.CS5, 10, 6, 1, 1)
        self.CS5.clicked.connect(lambda: self.matching("CS5")) 
        
        self.SB2 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.SB2.setObjectName("sb2")
        self.gridLayout.addWidget(self.SB2, 11, 2, 1, 1)
        self.SB2.clicked.connect(lambda: self.matching("SB2"))
        
        self.HT2 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HT2.setObjectName("ht2")
        self.gridLayout.addWidget(self.HT2, 11, 5, 1, 1)
        self.HT2.clicked.connect(lambda: self.matching("HT2"))
        
        self.HT3 = QtWidgets.QPushButton(self.gridLayoutWidget)
        self.HT3.setObjectName("ht3")
        self.gridLayout.addWidget(self.HT3, 11, 6, 1, 1)
        self.HT3.clicked.connect(lambda: self.matching("HT3"))
        
        self.detailEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.detailEdit.setGeometry(QtCore.QRect(300, 630, 491, 21))
        self.detailEdit.setObjectName("detailEdit")
        self.detailEdit.cursorPositionChanged.connect(self.display_editing)
        self.editingLabel = QtWidgets.QLabel(self.centralwidget)
        self.editingLabel.setGeometry(QtCore.QRect(300, 650, 491, 21))
        self.editingLabel.setObjectName("editingLabel")
        self.detailEdit.returnPressed.connect(self.edited)
        self.yearCombo = QtWidgets.QComboBox(self.centralwidget)
        self.yearCombo.setGeometry(QtCore.QRect(600, 50, 75, 21))
        self.yearCombo.setObjectName("yearCombo")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.yearCombo.addItem("")
        self.monthCombo = QtWidgets.QComboBox(self.centralwidget)
        self.monthCombo.setGeometry(QtCore.QRect(680, 50, 75, 21))
        self.monthCombo.setObjectName("monthCombo")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.monthCombo.addItem("")
        self.yearMonth = QtWidgets.QPushButton(self.centralwidget)
        self.yearMonth.setGeometry(QtCore.QRect(760, 48, 75, 25))
        self.yearMonth.setObjectName("yearMonth")
        self.yearMonth.clicked.connect(self.get_year_month)
        self.reportLabel = QtWidgets.QLabel(self.centralwidget)
        self.reportLabel.setGeometry(QtCore.QRect(40, 90, 191, 16))
        self.reportLabel.setObjectName("reportLabel")
        self.editLabel = QtWidgets.QLabel(self.centralwidget)
        self.editLabel.setGeometry(QtCore.QRect(300, 610, 291, 16))
        self.editLabel.setObjectName("editLabel")
        self.completeEdit = QtWidgets.QPushButton(self.centralwidget)
        self.completeEdit.setGeometry(QtCore.QRect(800, 630, 75, 21)) 
        self.completeEdit.setObjectName("completeEdit")
        self.completeEdit.clicked.connect(self.edited)
        
        self.save = QtWidgets.QPushButton(self.centralwidget)
        self.save.setGeometry(QtCore.QRect(530, 710, 80, 23)) 
        self.save.setObjectName("wrting_in")
        self.save.clicked.connect(self.writing_in)
        
        # self.rollback_e = QtWidgets.QPushButton(self.centralwidget)
        # self.rollback_e.setGeometry(QtCore.QRect(880, 630, 75, 21)) 
        # self.rollback_e.setObjectName("rollback_E")
        # self.rollback_e.clicked.connect(self.rollback_E)
        
        self.versionManage = QtWidgets.QLabel(self.centralwidget)
        self.versionManage.setGeometry(QtCore.QRect(980, 790, 121, 16))
        self.versionManage.setObjectName("versionManage")
        self.reportList = QtWidgets.QListWidget(self.centralwidget)
        self.reportList.setGeometry(QtCore.QRect(30, 110, 531, 431))
        self.reportList.setObjectName("reportList")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1108, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)
        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)
        self.reportList.setVerticalScrollBarPolicy(QtCore.Qt.ScrollBarAlwaysOff)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "설비이력카드 업데이트"))
        self.MP2002.setText(_translate("MainWindow", "MP2002"))
        self.MP1103.setText(_translate("MainWindow", "MP1103"))
        self.MP1101.setText(_translate("MainWindow", "MP1101"))
        self.MP1102.setText(_translate("MainWindow", "MP1102"))
        self.MP2001.setText(_translate("MainWindow", "MP2001"))
        self.MP1104.setText(_translate("MainWindow", "MP1104"))
        self.MP4005.setText(_translate("MainWindow", "MP4005"))
        self.MP4004.setText(_translate("MainWindow", "MP4004"))
        self.MP4003.setText(_translate("MainWindow", "MP4003"))
        self.MP4002.setText(_translate("MainWindow", "MP4002"))
        self.MP4001.setText(_translate("MainWindow", "MP4001"))
        self.MP2501.setText(_translate("MainWindow", "MP2501"))
        self.MP6304.setText(_translate("MainWindow", "MP6304"))
        self.MP6303.setText(_translate("MainWindow", "MP6303"))
        self.MP6302.setText(_translate("MainWindow", "MP6302"))
        self.MP6301.setText(_translate("MainWindow", "MP6301"))
        self.MP600.setText(_translate("MainWindow", "MP600"))
        self.MP4006.setText(_translate("MainWindow", "MP4006"))
        self.MP1000.setText(_translate("MainWindow", "MP1000"))
        self.MP650.setText(_translate("MainWindow", "MP650"))
        self.MP800.setText(_translate("MainWindow", "MP800"))
        self.HP201.setText(_translate("MainWindow", "HP201"))
        self.HP202.setText(_translate("MainWindow", "HP202"))
        self.HP203.setText(_translate("MainWindow", "HP203"))
        self.HP204.setText(_translate("MainWindow", "HP204"))
        self.HP205.setText(_translate("MainWindow", "HP205"))
        self.HP206.setText(_translate("MainWindow", "HP206"))
        self.HP207.setText(_translate("MainWindow", "HP207"))
        self.HP208.setText(_translate("MainWindow", "HP208"))
        self.SB1.setText(_translate("MainWindow", "SB1"))
        self.SB3.setText(_translate("MainWindow", "SB3"))
        self.HT1.setText(_translate("MainWindow", "HT1"))
        self.HP209.setText(_translate("MainWindow", "HP209"))
        self.CS1.setText(_translate("MainWindow", "CS1"))
        self.CS2.setText(_translate("MainWindow", "CS2"))
        self.CS3.setText(_translate("MainWindow", "CS3"))
        self.CS4.setText(_translate("MainWindow", "CS4"))
        self.CS5.setText(_translate("MainWindow", "CS5"))
        self.SB2.setText(_translate("MainWindow", "SB2"))
        self.HT2.setText(_translate("MainWindow", "HT2"))
        self.HT3.setText(_translate("MainWindow", "HT3"))
        self.adding.setText(_translate("MainWindow", "내용 추가"))
        self.deleting.setText(_translate("MainWindow", "지우기"))
        self.yearCombo.setItemText(0, _translate("MainWindow", "2021"))
        self.yearCombo.setItemText(1, _translate("MainWindow", "2022"))
        self.yearCombo.setItemText(2, _translate("MainWindow", "2023"))
        self.yearCombo.setItemText(3, _translate("MainWindow", "2024"))
        self.yearCombo.setItemText(4, _translate("MainWindow", "2025"))
        self.yearCombo.setItemText(5, _translate("MainWindow", "2026"))
        self.yearCombo.setItemText(6, _translate("MainWindow", "2027"))
        self.yearCombo.setItemText(7, _translate("MainWindow", "2028"))
        self.yearCombo.setItemText(8, _translate("MainWindow", "2029"))
        self.yearCombo.setItemText(9, _translate("MainWindow", "2030"))
        self.yearCombo.setItemText(10, _translate("MainWindow", "2031"))
        self.monthCombo.setItemText(0, _translate("MainWindow", "01"))
        self.monthCombo.setItemText(1, _translate("MainWindow", "02"))
        self.monthCombo.setItemText(2, _translate("MainWindow", "03"))
        self.monthCombo.setItemText(3, _translate("MainWindow", "04"))
        self.monthCombo.setItemText(4, _translate("MainWindow", "05"))
        self.monthCombo.setItemText(5, _translate("MainWindow", "06"))
        self.monthCombo.setItemText(6, _translate("MainWindow", "07"))
        self.monthCombo.setItemText(7, _translate("MainWindow", "08"))
        self.monthCombo.setItemText(8, _translate("MainWindow", "09"))
        self.monthCombo.setItemText(9, _translate("MainWindow", "10"))
        self.monthCombo.setItemText(10, _translate("MainWindow", "11"))
        self.monthCombo.setItemText(11, _translate("MainWindow", "12"))
        self.yearMonth.setText(_translate("MainWindow", "연월 설정"))
        self.reportLabel.setText(_translate("MainWindow", "생산기술팀 업무보고 목록"))
        self.editLabel.setText(_translate("MainWindow", "내용 편집"))
        self.completeEdit.setText(_translate("MainWindow", "편집 완료"))
        self.save.setText(_translate("MainWindow", "저장하기"))
        # self.rollback_e.setText(_translate("MainWindow", "되돌리기"))
        self.versionManage.setText(_translate("MainWindow", "Verson 1.1.3")) 
        # 1.0.1 매칭-딜리트-편집완료시 인덱스 바뀌는현상, 날짜 순서대로 정렬되게끔 수정
        # 1.0.2 CS5호기 버튼 추가
        # 1.0.3 스트링으로 정렬시 사전편찬순으로 정리 되던 것 정수형으로 정렬하여 순서 맞춤
        # 1.1.3 되돌리기 기능 존ㄴ ㅏ추가 

        self.deleting.setShortcut(_translate("MainWindow", "Del"))


    def inserting_records(self, records):

        for _ in reversed(records):
            self.reportList.addItem(_[1])

    
    def get_year_month(self):
        self.year = self.yearCombo.currentText()
        self.month = self.monthCombo.currentText()
        self.show_popup("입력 완료", "연월 입력 완료")

    def matching(self, machineName):

        if int(self.year) > 0:
            self.machineName = machineName
            self.idx = self.reportList.currentRow()
            self.matchingText = self.reportList.currentItem().text()
            self.detailEdit.setText(self.matchingText)  # 현재 로우 내용 가져올 것 
            self.detailEdit.setFocus()
            self.detailEdit.setCursorPosition(0)
        elif int(self.year) == 0:
            self.show_popup("오류", "날짜를 먼저 설정 해주세요.")
            self.detailEdit.clear()
        self.matching_sig = 1


    def edited(self):
        if self.matching_sig == True:
            date = int(self.reversedRecords[self.idx][0])
            uneditedDetail = self.reversedRecords[self.idx][1]

            print(self.reversedRecords[self.idx][1])
            print(self.editedRec)

            if self.machineName in self.matched.keys():
                self.matched[self.machineName].append([date, self.editedRec])

            elif not self.machineName in self.matched.keys():
                self.matched[self.machineName] = [[date, self.editedRec]]

            self.machineNameList.append(self.machineName)
            self.deletedEdited.append([self.machineName, self.idx, date, uneditedDetail, self.editedRec])
            self.reportList.takeItem(self.idx)
            self.reversedRecords.pop(self.idx)
            self.detailEdit.clear()
            self.matching_sig = 0
            self.reportList.setFocus()
            print(self.deletedEdited)
            # print(self.matched)
            self.listOrEdit = 0
        elif self.matching_sig == False:
            self.show_popup('오류', '설비를 선택 해주세요.')
        

    def show_popup(self, title, message):
        msg = QtWidgets.QMessageBox()
        msg.setWindowTitle(title)
        msg.setText(message)
        msg.setIcon(QtWidgets.QMessageBox.Warning)
        if title == "저장 완료":
            msg.setIcon(QtWidgets.QMessageBox.Information)
        if title == '입력 완료':
            msg.setIcon(QtWidgets.QMessageBox.Information)
        msg.exec_()
        
    
    def adding_details(self):
        try:
            self.idx = self.reportList.currentRow()
            self.matched[self.machineName][-1].append(self.reversedRecords[self.idx][1])
            self.reportList.takeItem(self.idx)
            self.reversedRecords.pop(self.idx)
        except KeyError:
            self.show_popup("오류", "내용 추가할 설비를 선택해주세요.")
    
    
    def deleting_details(self):
        text = self.reportList.currentItem().text()
        idx = self.reportList.currentRow()
        self.deletedLists.append([str(idx), self.reversedRecords[idx][0], text])
        self.reportList.takeItem(idx)
        self.reversedRecords.pop(idx)
        self.listOrEdit = 1



    def rollback(self):
        
        if not self.deletedEdited and not self.deletedLists:
            self.show_popup("오류", "되돌릴 내용이 없습니다.")
            self.listOrEdit == 3
            
        else:

            if self.listOrEdit == 0:

                Idx = int(self.deletedEdited[-1][-4])

                date = self.deletedEdited[-1][-3]

                machineName = self.deletedEdited[-1][0]

                uneditedDetail = self.deletedEdited[-1][-2]

                del self.matched[machineName][-1]

                self.reportList.insertItem(Idx, uneditedDetail)
                self.reversedRecords.insert(Idx, [date, uneditedDetail])

                self.reportList.setCurrentRow(Idx)

                self.detailEdit.setText(uneditedDetail)

                self.machineName = self.deletedEdited[-1][0]
                if len(self.deletedEdited) == 1 and self.deletedLists:
                    self.listOrEdit == 1
                        
                self.deletedEdited.pop(-1)

                self.matching_sig = 1

                self.machineNameList.pop(-1)
            
               
                    
                
                    
            if self.listOrEdit == 1:
                    setRow = int(self.deletedLists[-1][0])
                    self.reportList.insertItem(int(self.deletedLists[-1][0]), self.deletedLists[-1][2])
                    self.reversedRecords.insert(int(self.deletedLists[-1][0]), [self.deletedLists[-1][1], self.deletedLists[-1][2]])
                    if len(self.deletedLists) == 1 and self.deletedEdited:
                        self.listOrEdit == 0
                    self.deletedLists.pop(-1)
                    self.reportList.setCurrentRow(setRow)
                    
    
        
    # def rollback_E(self):
    #     try:
    #         setRow = int(self.deletedEdited[-1][-4])
    #         self.reportList.insertItem(int(self.deletedEdited[-1][-4]), self.deletedEdited[-1][-2])
    #         self.reportList.setCurrentRow(setRow)
    #         self.detailEdit.setText(self.deletedEdited[-1][-2])
    #         self.detailEdit.setFocus()
    #         print(self.deletedEdited)
    #         self.reversedRecords.insert(self.deletedEdited[-1][2], [self.deletedEdited[-1][-3], self.deletedEdited[-1][-2]])
    #         self.deletedEdited.pop(-1)
    #         self.detailEdit.setCursorPosition(0)
    #         self.matching_sig = 1
    #         del self.matched[self.machineNameList[-1]][-1]
    #         self.machineName = self.machineNameList[-1]
    #         self.machineNameList.pop(-1)
            
    #     except IndexError:
    #         self.show_popup("오류", "되돌릴 내용이 없습니다.")
            
    # def rollback_L(self):
    #         try:
    #             setRow = int(self.deletedLists[-1][0])
    #             self.reportList.insertItem(int(self.deletedLists[-1][0]), self.deletedLists[-1][1])
    #             self.reversedRecords.append(self.deletedLists[-1])
    #             self.deletedLists.pop(-1)
    #             self.reportList.setCurrentRow(setRow)
                
    #         except IndexError:
    #             self.show_popup("오류", "되돌릴 내용이 없습니다.")
        

    def display_editing(self, old, new):
        oldText = self.detailEdit.text()
        self.editingLabel.setText(oldText[new:])
        self.editedRec = oldText[new:]
    


    def finding_empty(self, sheet, columnLetter):
        cells = []
        for cell in sheet[columnLetter]: 
            if cell.value == None and cell.row > 41:
                cells.append(str(cell.row))
                cells.append(str(self.columns[columnLetter]))
                break

        return cells


    def writing_in(self):
        csvColumns = []
        for machine in self.matched:
            self.matched[machine].sort()

        for workbook in self.workbookList:
            
            wb = load_workbook(workbook)

            for machine in self.matched:
                try:
                    ws = wb[machine]
                except KeyError:
                    continue
                
                try:
                    for i in range(100):
                        if len(self.matched[machine][0]) == 2:
                            if len(str(self.matched[machine][0][0])) == 1:
                                self.matched[machine][0][0] = "0" + str(self.matched[machine][0][0])

                                cells = self.finding_empty(ws, 'C')
                                ws.cell(row=int(cells[0]), column=int(self.columns["C"])).value = self.matched[machine][0][1]
                                ws.cell(row=int(cells[0]), column=int(self.columns["A"])).value = datetime.datetime.strptime(str(self.year) + str(self.month) + str(self.matched[machine][0][0]), self.dateFormat).date()
                                del self.matched[machine][0]
                            elif len(str(self.matched[machine][0][0])) == 2:

                                cells = self.finding_empty(ws, 'C')
                                ws.cell(row=int(cells[0]), column=int(self.columns["C"])).value = self.matched[machine][0][1]
                                ws.cell(row=int(cells[0]), column=int(self.columns["A"])).value = datetime.datetime.strptime(str(self.year) + str(self.month) + str(self.matched[machine][0][0]), self.dateFormat).date()
                                del self.matched[machine][0]

                        elif len(self.matched[machine][0]) == 3:
                            if len(str(self.matched[machine][0][0])) == 1:
                                cells = self.finding_empty(ws, 'C')
                                ws.cell(row=int(cells[0]), column=int(self.columns["C"])).value = self.matched[machine][0][1]
                                ws.cell(row=int(cells[0]), column=int(self.columns["A"])).value = datetime.datetime.strptime(str(self.year) + str(self.month) + str(self.matched[machine][0][0]), self.dateFormat).date()
                                cells = self.finding_empty(ws, 'C')
                                ws.cell(row=int(cells[0]), column=int(self.columns["C"])).value = self.matched[machine][0][2]
                                del self.matched[machine][0]
                            
                            elif len(str(self.matched[machine][0][0])) == 2:
                                cells = self.finding_empty(ws, 'C')
                                ws.cell(row=int(cells[0]), column=int(self.columns["C"])).value = self.matched[machine][0][1]
                                ws.cell(row=int(cells[0]), column=int(self.columns["A"])).value = datetime.datetime.strptime(str(self.year) + str(self.month) + str(self.matched[machine][0][0]), self.dateFormat).date()
                                cells = self.finding_empty(ws, 'C')
                                ws.cell(row=int(cells[0]), column=int(self.columns["C"])).value = self.matched[machine][0][2]
                                del self.matched[machine][0]

                        if len(self.matched[machine][0]) == 0:
                            del self.matched[machine]
                except IndexError:
                    continue
                      
            wb.save('{} 설비이력카드 (~{}.{}.) .xlsx'.format(str(workbook)[:-5], self.year, self.month))
            wb.close()
        
        self.show_popup("저장 완료", "저장 완료됨")
         



if __name__ == "__main__":
    import sys
    records = importing_recrods('files/개발업무.xlsx')
    app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet(open('style.css').read())
    MainWindow = QtWidgets.QMainWindow()
    ui = Ui_MainWindow()
    ui.setupUi(MainWindow)
    ui.inserting_records(records)
    MainWindow.show()
    sys.exit(app.exec())